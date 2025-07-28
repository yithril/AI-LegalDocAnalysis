import aiofiles
import time
from typing import AsyncIterator
from pathlib import Path
from models.tenant.document import DocumentStatus
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..text_extraction_strategy import TextExtractionStrategy
from ...models.extraction_result import TextExtractionResult


async def _empty_iterator() -> AsyncIterator[str]:
    """Return an empty async iterator."""
    if False:  # This will never be True, but satisfies the type checker
        yield ""


class ExcelStrategy(TextExtractionStrategy):
    async def can_handle(self, file_path: str, mime_type: str) -> bool:
        path = Path(file_path)
        return (
            path.suffix.lower() in self.get_supported_extensions() or
            mime_type in self.get_supported_mime_types()
        )

    async def validate_file(self, file_path: str) -> bool:
        """Validate Excel file by attempting to load it with openpyxl."""
        try:
            path = Path(file_path)

            if not path.exists():
                return False
            
            if not path.is_file():
                return False
            
            if await self.is_empty_file(file_path):
                return True  # Empty Excel files are valid
            
            # Try to load the Excel file with openpyxl
            try:
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                # Check if there are any sheets
                if len(workbook.sheetnames) == 0:
                    return False
                
                # Try to access the first sheet to validate
                sheet = workbook[workbook.sheetnames[0]]
                # Just check if we can access the sheet - don't read all data
                _ = sheet.max_row
                _ = sheet.max_column
                
                workbook.close()
                return True
            except (InvalidFileException, Exception):
                return False
            
        except Exception:
            return False

    async def extract_text_from_stream(self, file_path: str) -> TextExtractionResult:
        start_time = time.time()
        
        try:
            # Basic file validation
            path = Path(file_path)
            if not path.exists() or not path.is_file():
                processing_time = time.time() - start_time
                return TextExtractionResult.failure_result(
                    status=DocumentStatus.CORRUPTED,
                    file_path=file_path,
                    strategy_used=self.__class__.__name__,
                    error_message=f"Invalid file or corrupted Excel file: {file_path}",
                    processing_time=processing_time
                )
            
            if await self.is_empty_file(file_path):
                processing_time = time.time() - start_time
                return TextExtractionResult.success_result(
                    text_chunks=_empty_iterator(),
                    file_path=file_path,
                    strategy_used=self.__class__.__name__,
                    processing_time=processing_time,
                    metadata={"file_size": 0, "format": "excel", "sheets": 0, "total_rows": 0}
                )

            # Validate Excel structure
            if not await self.validate_file(file_path):
                processing_time = time.time() - start_time
                return TextExtractionResult.failure_result(
                    status=DocumentStatus.CORRUPTED,
                    file_path=file_path,
                    strategy_used=self.__class__.__name__,
                    error_message=f"Invalid file or corrupted Excel file: {file_path}",
                    processing_time=processing_time
                )

            async def text_chunks():
                try:
                    # Load the workbook
                    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                    chunk_size = 8192
                    current_chunk = ""
                    total_rows = 0
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        # Add sheet header
                        sheet_header = f"\n=== Sheet: {sheet_name} ===\n"
                        if len(current_chunk) + len(sheet_header) > chunk_size:
                            if current_chunk:
                                yield current_chunk
                            current_chunk = sheet_header
                        else:
                            current_chunk += sheet_header
                        
                        # Process each row in the sheet
                        for row in sheet.iter_rows(values_only=True):
                            total_rows += 1
                            
                            # Convert row to readable text (filter out None values)
                            row_values = [str(value) if value is not None else "" for value in row]
                            row_text = " | ".join(row_values) + "\n"
                            
                            if len(current_chunk) + len(row_text) > chunk_size:
                                if current_chunk:
                                    yield current_chunk
                                current_chunk = row_text
                            else:
                                current_chunk += row_text
                    
                    # Yield remaining content
                    if current_chunk:
                        yield current_chunk
                    
                    workbook.close()
                        
                except Exception as e:
                    # If Excel parsing fails, yield error message
                    yield f"Error parsing Excel file: {str(e)}"

            # Get metadata
            metadata = {
                "file_size": path.stat().st_size,
                "format": "excel"
            }
            
            # Try to get sheet and row count for metadata
            try:
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet_count = len(workbook.sheetnames)
                total_rows = 0
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_rows += sheet.max_row
                
                metadata.update({
                    "sheets": sheet_count,
                    "total_rows": total_rows,
                    "sheet_names": workbook.sheetnames
                })
                
                workbook.close()
            except:
                pass
            
            processing_time = time.time() - start_time
            
            return TextExtractionResult.success_result(
                text_chunks=text_chunks(),
                file_path=file_path,
                strategy_used=self.__class__.__name__,
                processing_time=processing_time,
                metadata=metadata
            )
        except Exception as e:
            processing_time = time.time() - start_time
            return TextExtractionResult.failure_result(
                status=DocumentStatus.EXTRACTION_FAILED,
                file_path=file_path,
                strategy_used=self.__class__.__name__,
                error_message=str(e),
                processing_time=processing_time
            )

    def get_supported_extensions(self) -> list[str]:
        return ['.xlsx', '.xls']

    def get_supported_mime_types(self) -> list[str]:
        return [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel'  # .xls
        ] 