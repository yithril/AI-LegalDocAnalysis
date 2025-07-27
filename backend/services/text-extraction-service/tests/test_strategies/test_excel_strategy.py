import pytest
import tempfile
import os
from pathlib import Path
from unittest.mock import AsyncMock, patch, MagicMock
from common_lib.document_enums import DocumentStatus
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from text_extraction_service.strategies.text_strategies.excel_strategy import ExcelStrategy


class TestExcelStrategy:
    @pytest.fixture
    def strategy(self):
        return ExcelStrategy()

    @pytest.fixture
    def simple_excel_file(self):
        """Create a simple Excel file for testing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add headers
        ws['A1'] = 'Name'
        ws['B1'] = 'Age'
        ws['C1'] = 'City'
        
        # Add data
        ws['A2'] = 'John Doe'
        ws['B2'] = 30
        ws['C2'] = 'New York'
        ws['A3'] = 'Jane Smith'
        ws['B3'] = 25
        ws['C3'] = 'Los Angeles'
        ws['A4'] = 'Bob Johnson'
        ws['B4'] = 35
        ws['C4'] = 'Chicago'
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        yield temp_path
        os.unlink(temp_path)

    @pytest.fixture
    def multi_sheet_excel_file(self):
        """Create an Excel file with multiple sheets."""
        wb = Workbook()
        
        # First sheet
        ws1 = wb.active
        ws1.title = "Employees"
        ws1['A1'] = 'ID'
        ws1['B1'] = 'Name'
        ws1['C1'] = 'Department'
        ws1['A2'] = 1
        ws1['B2'] = 'John Doe'
        ws1['C2'] = 'Engineering'
        ws1['A3'] = 2
        ws1['B3'] = 'Jane Smith'
        ws1['C3'] = 'Marketing'
        
        # Second sheet
        ws2 = wb.create_sheet("Projects")
        ws2['A1'] = 'Project ID'
        ws2['B1'] = 'Project Name'
        ws2['C1'] = 'Status'
        ws2['A2'] = 'P001'
        ws2['B2'] = 'Website Redesign'
        ws2['C2'] = 'Active'
        ws2['A3'] = 'P002'
        ws2['B3'] = 'Mobile App'
        ws2['C3'] = 'Completed'
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        yield temp_path
        os.unlink(temp_path)

    @pytest.fixture
    def legal_excel_file(self):
        """Create a legal-themed Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cases"
        
        # Add headers
        ws['A1'] = 'Case ID'
        ws['B1'] = 'Client Name'
        ws['C1'] = 'Case Type'
        ws['D1'] = 'Status'
        ws['E1'] = 'Filing Date'
        
        # Add data
        ws['A2'] = 'CASE-2024-001'
        ws['B2'] = 'John Doe'
        ws['C2'] = 'Contract Dispute'
        ws['D2'] = 'Active'
        ws['E2'] = '2024-01-15'
        ws['A3'] = 'CASE-2024-002'
        ws['B3'] = 'Jane Smith'
        ws['C3'] = 'Employment Law'
        ws['D3'] = 'Pending'
        ws['E3'] = '2024-01-20'
        ws['A4'] = 'CASE-2024-003'
        ws['B4'] = 'Bob Johnson'
        ws['C4'] = 'Personal Injury'
        ws['D4'] = 'Closed'
        ws['E4'] = '2024-01-25'
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        yield temp_path
        os.unlink(temp_path)

    @pytest.fixture
    def empty_excel_file(self):
        """Create an empty Excel file."""
        wb = Workbook()
        ws = wb.active
        # Remove all content
        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        yield temp_path
        os.unlink(temp_path)

    @pytest.fixture
    def invalid_excel_file(self):
        """Create an invalid Excel file by creating a text file with .xlsx extension."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(b'This is not an Excel file')
            temp_path = f.name
        yield temp_path
        os.unlink(temp_path)

    @pytest.mark.asyncio
    async def test_can_handle_with_excel_extension(self, strategy):
        """Test can_handle returns True for Excel file extensions."""
        assert await strategy.can_handle("data.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") is True
        assert await strategy.can_handle("data.xls", "application/vnd.ms-excel") is True
        assert await strategy.can_handle("export.XLSX", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") is True

    @pytest.mark.asyncio
    async def test_can_handle_with_excel_mime_type(self, strategy):
        """Test can_handle returns True for Excel MIME types."""
        assert await strategy.can_handle("data.unknown", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") is True
        assert await strategy.can_handle("data.unknown", "application/vnd.ms-excel") is True

    @pytest.mark.asyncio
    async def test_can_handle_with_unsupported_file(self, strategy):
        """Test can_handle returns False for unsupported files."""
        assert await strategy.can_handle("test.txt", "text/plain") is False
        assert await strategy.can_handle("test.csv", "text/csv") is False
        assert await strategy.can_handle("test.pdf", "application/pdf") is False
        assert await strategy.can_handle("test.docx", "application/vnd.openxmlformats-officedocument.wordprocessingml.document") is False

    @pytest.mark.asyncio
    async def test_validate_file_with_valid_excel(self, strategy, simple_excel_file):
        """Test validate_file returns True for valid Excel files."""
        assert await strategy.validate_file(simple_excel_file) is True

    @pytest.mark.asyncio
    async def test_validate_file_with_invalid_excel(self, strategy, invalid_excel_file):
        """Test validate_file returns False for invalid Excel files."""
        assert await strategy.validate_file(invalid_excel_file) is False

    @pytest.mark.asyncio
    async def test_validate_file_with_empty_excel(self, strategy, empty_excel_file):
        """Test validate_file returns True for empty Excel files."""
        assert await strategy.validate_file(empty_excel_file) is True

    @pytest.mark.asyncio
    async def test_validate_file_with_nonexistent_file(self, strategy):
        """Test validate_file returns False for nonexistent files."""
        assert await strategy.validate_file("nonexistent.xlsx") is False

    @pytest.mark.asyncio
    async def test_validate_file_with_directory(self, strategy):
        """Test validate_file returns False for directories."""
        with tempfile.TemporaryDirectory() as temp_dir:
            assert await strategy.validate_file(temp_dir) is False

    @pytest.mark.asyncio
    async def test_extract_text_from_stream_with_simple_excel(self, strategy, simple_excel_file):
        """Test extracting text from a simple Excel file."""
        result = await strategy.extract_text_from_stream(simple_excel_file)
        
        assert result.success is True
        assert result.status == DocumentStatus.PROCESSED
        assert result.strategy_used == "ExcelStrategy"
        assert result.file_path == simple_excel_file
        assert result.processing_time is not None
        assert result.processing_time >= 0  # Allow 0 for very fast processing
        
        # Check metadata
        assert "file_size" in result.metadata
        assert "format" in result.metadata
        assert result.metadata["format"] == "excel"
        assert "sheets" in result.metadata
        assert result.metadata["sheets"] == 1
        
        # Check content
        content = await result.get_text_content()
        assert "=== Sheet: Sheet1 ===" in content
        assert "Name | Age | City" in content
        assert "John Doe | 30 | New York" in content
        assert "Jane Smith | 25 | Los Angeles" in content

    @pytest.mark.asyncio
    async def test_extract_text_from_stream_with_multi_sheet_excel(self, strategy, multi_sheet_excel_file):
        """Test extracting text from an Excel file with multiple sheets."""
        result = await strategy.extract_text_from_stream(multi_sheet_excel_file)
        
        assert result.success is True
        assert result.status == DocumentStatus.PROCESSED
        
        # Check metadata
        assert "sheets" in result.metadata
        assert result.metadata["sheets"] == 2
        assert "sheet_names" in result.metadata
        assert "Employees" in result.metadata["sheet_names"]
        assert "Projects" in result.metadata["sheet_names"]
        
        # Check content
        content = await result.get_text_content()
        assert "=== Sheet: Employees ===" in content
        assert "=== Sheet: Projects ===" in content
        assert "ID | Name | Department" in content
        assert "Project ID | Project Name | Status" in content

    @pytest.mark.asyncio
    async def test_extract_text_from_stream_with_legal_excel(self, strategy, legal_excel_file):
        """Test extracting text from a legal-themed Excel file."""
        result = await strategy.extract_text_from_stream(legal_excel_file)
        
        assert result.success is True
        assert result.status == DocumentStatus.PROCESSED
        
        # Check content
        content = await result.get_text_content()
        assert "=== Sheet: Cases ===" in content
        assert "Case ID | Client Name | Case Type | Status | Filing Date" in content
        assert "CASE-2024-001 | John Doe | Contract Dispute | Active | 2024-01-15" in content
        assert "CASE-2024-002 | Jane Smith | Employment Law | Pending | 2024-01-20" in content

    @pytest.mark.asyncio
    async def test_extract_text_from_stream_with_invalid_excel(self, strategy, invalid_excel_file):
        """Test extracting text from an invalid Excel file."""
        result = await strategy.extract_text_from_stream(invalid_excel_file)
        
        assert result.success is False
        assert result.status == DocumentStatus.CORRUPTED
        assert "Invalid file or corrupted Excel file" in result.error_message

    @pytest.mark.asyncio
    async def test_extract_text_from_stream_with_empty_excel(self, strategy, empty_excel_file):
        """Test extracting text from an empty Excel file."""
        result = await strategy.extract_text_from_stream(empty_excel_file)
        
        assert result.success is True
        assert result.status == DocumentStatus.PROCESSED
        
        # Check metadata
        assert result.metadata["sheets"] == 1
        # Even empty Excel files have at least 1 row due to sheet metadata
        assert result.metadata["total_rows"] >= 0
        
        # Check content - empty Excel files still have sheet headers
        content = await result.get_text_content()
        assert "=== Sheet:" in content  # Should have sheet header
        # But no actual data rows
        lines = content.strip().split('\n')
        assert len(lines) <= 2  # Just sheet header and maybe empty line

    @pytest.mark.asyncio
    async def test_extract_text_from_stream_with_nonexistent_file(self, strategy):
        """Test extracting text from a nonexistent file."""
        result = await strategy.extract_text_from_stream("nonexistent.xlsx")
        
        assert result.success is False
        assert result.status == DocumentStatus.CORRUPTED
        assert "Invalid file or corrupted Excel file" in result.error_message

    @pytest.mark.asyncio
    async def test_extract_text_from_stream_chunking(self, strategy, legal_excel_file):
        """Test that text extraction properly chunks large content."""
        result = await strategy.extract_text_from_stream(legal_excel_file)
        
        assert result.success is True
        
        # Test chunking by iterating through chunks
        chunks = []
        async for chunk in result.text_chunks:
            chunks.append(chunk)
        
        # Should have at least one chunk
        assert len(chunks) > 0
        
        # All chunks should contain some content
        for chunk in chunks:
            assert len(chunk) > 0
        
        # Combined content should contain expected data
        combined_content = "".join(chunks)
        assert "=== Sheet: Cases ===" in combined_content
        assert "Case ID | Client Name | Case Type | Status | Filing Date" in combined_content

    def test_get_supported_extensions(self, strategy):
        """Test get_supported_extensions returns correct extensions."""
        expected_extensions = ['.xlsx', '.xls']
        assert strategy.get_supported_extensions() == expected_extensions

    def test_get_supported_mime_types(self, strategy):
        """Test get_supported_mime_types returns correct MIME types."""
        expected_mime_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel'
        ]
        assert strategy.get_supported_mime_types() == expected_mime_types

    @pytest.mark.asyncio
    async def test_extract_text_preserves_structure(self, strategy, simple_excel_file):
        """Test that Excel structure is preserved in extracted text."""
        result = await strategy.extract_text_from_stream(simple_excel_file)
        
        assert result.success is True
        
        content = await result.get_text_content()
        
        # Check that sheet headers are preserved
        assert "=== Sheet: Sheet1 ===" in content
        
        # Check that data rows are preserved
        lines = content.strip().split('\n')
        assert len(lines) >= 4  # Sheet header + header row + 3 data rows
        
        # Check that each line has the expected number of fields
        for line in lines:
            if line.strip() and not line.startswith("==="):  # Skip sheet headers
                fields = line.split(' | ')
                assert len(fields) == 3

    @pytest.mark.asyncio
    async def test_metadata_includes_sheet_info(self, strategy, multi_sheet_excel_file):
        """Test that metadata includes sheet information."""
        result = await strategy.extract_text_from_stream(multi_sheet_excel_file)
        
        assert result.success is True
        assert "sheets" in result.metadata
        assert result.metadata["sheets"] == 2
        assert "sheet_names" in result.metadata
        assert len(result.metadata["sheet_names"]) == 2
        assert "Employees" in result.metadata["sheet_names"]
        assert "Projects" in result.metadata["sheet_names"]
        assert "total_rows" in result.metadata
        assert result.metadata["total_rows"] > 0 